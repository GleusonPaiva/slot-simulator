"""
=============================================================
  SLOT MACHINE SIMULATOR - Game Math Designer Portfolio
=============================================================
  Projeto de portfólio para mercado de iGaming
  Simula um slot de 5 rolos, calcula RTP teórico e simulado,
  gera relatório em Excel e PDF automaticamente.
=============================================================
"""

import random
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import os
import datetime

# ─────────────────────────────────────────────
#  1. CONFIGURAÇÃO DO JOGO (MATH MODEL)
# ─────────────────────────────────────────────

NOME_DO_JOGO = "Golden Reels"
VERSAO = "1.0.0"
RTP_ALVO = 96.0  # % — meta de retorno ao jogador

# Símbolos do jogo e seus PESOS em cada rolo
# Peso maior = aparece mais vezes = paga menos
# Peso menor = aparece menos vezes = paga mais
SIMBOLOS = {
    "🍒 Cereja":    {"peso": 30, "cor": "#e74c3c"},
    "🍋 Limão":     {"peso": 25, "cor": "#f1c40f"},
    "🍊 Laranja":   {"peso": 20, "cor": "#e67e22"},
    "⭐ Estrela":   {"peso": 15, "cor": "#9b59b6"},
    "💎 Diamante":  {"peso": 7,  "cor": "#3498db"},
    "7️⃣  Sete":      {"peso": 3,  "cor": "#e74c3c"},
}

# Tabela de pagamentos — quantas vezes a aposta o jogador recebe
# por ter N símbolos iguais na linha central (da esquerda para direita)
PAGAMENTOS = {
    #  símbolo          3x     4x     5x
    "🍒 Cereja":    {3: 2,   4: 5,   5: 10},
    "🍋 Limão":     {3: 3,   4: 8,   5: 15},
    "🍊 Laranja":   {3: 5,   4: 12,  5: 25},
    "⭐ Estrela":   {3: 10,  4: 25,  5: 50},
    "💎 Diamante":  {3: 25,  4: 75,  5: 200},
    "7️⃣  Sete":      {3: 50,  4: 150, 5: 500},
}

NUM_ROLOS = 5
APOSTA_BASE = 1.0  # valor padrão da aposta (em unidades)

# ─────────────────────────────────────────────
#  2. FUNÇÕES DO MOTOR DO SLOT
# ─────────────────────────────────────────────

def criar_rolo():
    """
    Cria a lista de símbolos de um rolo com base nos pesos.
    Peso 30 = 30 cópias do símbolo no rolo.
    Total de símbolos por rolo = soma de todos os pesos.
    """
    rolo = []
    for simbolo, dados in SIMBOLOS.items():
        rolo.extend([simbolo] * dados["peso"])
    return rolo

def girar(rolos):
    """
    Gira os 5 rolos e retorna o resultado central (linha de pagamento).
    Em produção real, cada rolo teria posições independentes.
    """
    return [random.choice(rolo) for rolo in rolos]

def calcular_premio(resultado, aposta=APOSTA_BASE):
    """
    Verifica o resultado e calcula o prêmio.
    Regra: conta quantos símbolos iguais aparecem
    da ESQUERDA para a DIREITA sem interrupção.
    """
    primeiro = resultado[0]
    contagem = 1

    for i in range(1, NUM_ROLOS):
        if resultado[i] == primeiro:
            contagem += 1
        else:
            break  # sequência interrompida

    if contagem >= 3 and primeiro in PAGAMENTOS:
        multiplicador = PAGAMENTOS[primeiro].get(contagem, 0)
        return aposta * multiplicador

    return 0.0

# ─────────────────────────────────────────────
#  3. CÁLCULO DO RTP TEÓRICO
# ─────────────────────────────────────────────

def calcular_rtp_teorico():
    """
    Calcula o RTP matematicamente, sem simulação.
    
    Fórmula:
    RTP = Σ (probabilidade de cada combinação × seu pagamento)
    
    A probabilidade de um símbolo em um rolo =
    peso_do_simbolo / soma_total_dos_pesos
    """
    total_pesos = sum(s["peso"] for s in SIMBOLOS.values())
    rtp_total = 0.0
    detalhes = []

    for simbolo, dados in SIMBOLOS.items():
        prob_por_rolo = dados["peso"] / total_pesos

        for qtd in [3, 4, 5]:
            if simbolo not in PAGAMENTOS:
                continue
            multiplicador = PAGAMENTOS[simbolo].get(qtd, 0)
            if multiplicador == 0:
                continue

            # Probabilidade de ter exatamente N símbolos seguidos
            # P = prob^N × (1 - prob)^(5-N) para posições após N
            # Simplificado: prob dos N primeiros × prob de NÃO ser o símbolo depois
            if qtd == 5:
                prob_combo = prob_por_rolo ** 5
            else:
                prob_nao = 1 - prob_por_rolo
                prob_combo = (prob_por_rolo ** qtd) * prob_nao

            contribuicao = prob_combo * multiplicador * 100  # em %
            rtp_total += contribuicao

            detalhes.append({
                "Símbolo": simbolo,
                "Quantidade": f"{qtd}x",
                "Multiplicador": f"{multiplicador}x",
                "Probabilidade": f"{prob_combo:.6f}",
                "Prob (%)": f"{prob_combo*100:.4f}%",
                "Contribuição RTP": f"{contribuicao:.4f}%"
            })

    return rtp_total, detalhes

# ─────────────────────────────────────────────
#  4. SIMULAÇÃO (MONTE CARLO)
# ─────────────────────────────────────────────

def simular(num_rodadas=10_000_000, aposta=APOSTA_BASE, seed=42):
    """
    Simula N rodadas do slot e coleta estatísticas.
    
    O método Monte Carlo é o padrão da indústria para
    validar math models antes da certificação.
    """
    print(f"\n⚙️  Iniciando simulação de {num_rodadas:,} rodadas...")
    print(f"    Isso pode levar alguns segundos...\n")

    random.seed(seed)  # seed garante reprodutibilidade dos resultados
    rolos = [criar_rolo() for _ in range(NUM_ROLOS)]

    total_apostado = 0.0
    total_pago = 0.0
    contagem_premios = defaultdict(int)
    historico_rtp = []       # para gráfico de convergência
    distribuicao_wins = []   # para gráfico de distribuição

    checkpoint = num_rodadas // 20  # registra RTP a cada 5%

    for i in range(num_rodadas):
        resultado = girar(rolos)
        premio = calcular_premio(resultado, aposta)

        total_apostado += aposta
        total_pago += premio

        if premio > 0:
            contagem_premios[premio] += 1
            distribuicao_wins.append(premio)

        # Registra snapshots do RTP ao longo do tempo
        if (i + 1) % checkpoint == 0:
            rtp_atual = (total_pago / total_apostado) * 100
            historico_rtp.append({
                "rodada": i + 1,
                "rtp": rtp_atual
            })
            progresso = int((i + 1) / num_rodadas * 20)
            barra = "█" * progresso + "░" * (20 - progresso)
            print(f"    [{barra}] {(i+1)/num_rodadas*100:.0f}% — RTP atual: {rtp_atual:.2f}%")

    rtp_final = (total_pago / total_apostado) * 100
    hit_rate = (sum(contagem_premios.values()) / num_rodadas) * 100

    print(f"\n✅ Simulação concluída!")
    print(f"   RTP Simulado:  {rtp_final:.2f}%")
    print(f"   RTP Alvo:      {RTP_ALVO:.2f}%")
    print(f"   Hit Rate:      {hit_rate:.2f}%")

    return {
        "rtp_final": rtp_final,
        "hit_rate": hit_rate,
        "total_apostado": total_apostado,
        "total_pago": total_pago,
        "num_rodadas": num_rodadas,
        "contagem_premios": dict(contagem_premios),
        "historico_rtp": historico_rtp,
        "distribuicao_wins": distribuicao_wins,
    }

# ─────────────────────────────────────────────
#  5. GERAÇÃO DE GRÁFICOS
# ─────────────────────────────────────────────

def gerar_graficos(resultados, rtp_teorico, output_dir):
    """Gera os gráficos que serão usados no relatório."""
    graficos = []
    plt.style.use('dark_background')

    # ── Gráfico 1: Convergência do RTP ──
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor('#1a1a2e')
    ax.set_facecolor('#16213e')

    historico = resultados["historico_rtp"]
    rodadas = [h["rodada"] for h in historico]
    rtps = [h["rtp"] for h in historico]

    ax.plot(rodadas, rtps, color='#00d4ff', linewidth=2, label='RTP Simulado')
    ax.axhline(y=rtp_teorico, color='#ffd700', linestyle='--', linewidth=1.5, label=f'RTP Teórico ({rtp_teorico:.2f}%)')
    ax.axhline(y=RTP_ALVO, color='#ff6b6b', linestyle=':', linewidth=1.5, label=f'RTP Alvo ({RTP_ALVO}%)')

    ax.set_xlabel('Número de Rodadas', color='white', fontsize=11)
    ax.set_ylabel('RTP (%)', color='white', fontsize=11)
    ax.set_title('Convergência do RTP — Lei dos Grandes Números', color='white', fontsize=13, fontweight='bold')
    ax.legend(facecolor='#1a1a2e', edgecolor='#444', labelcolor='white')
    ax.tick_params(colors='white')
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.1f}M'))
    ax.grid(color='#333', linestyle='--', alpha=0.5)

    caminho = os.path.join(output_dir, 'grafico_convergencia.png')
    plt.tight_layout()
    plt.savefig(caminho, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    graficos.append(caminho)

    # ── Gráfico 2: Pesos dos símbolos ──
    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor('#1a1a2e')
    ax.set_facecolor('#16213e')

    nomes = list(SIMBOLOS.keys())
    pesos = [s["peso"] for s in SIMBOLOS.values()]
    cores = [s["cor"] for s in SIMBOLOS.values()]
    total = sum(pesos)
    probs = [p/total*100 for p in pesos]

    bars = ax.barh(nomes, probs, color=cores, edgecolor='#333', height=0.6)
    for bar, prob in zip(bars, probs):
        ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height()/2,
                f'{prob:.1f}%', va='center', color='white', fontsize=10)

    ax.set_xlabel('Probabilidade por Rolo (%)', color='white', fontsize=11)
    ax.set_title('Probabilidade de Cada Símbolo por Rolo', color='white', fontsize=13, fontweight='bold')
    ax.tick_params(colors='white')
    ax.grid(axis='x', color='#333', linestyle='--', alpha=0.5)
    ax.set_xlim(0, max(probs) * 1.2)

    caminho = os.path.join(output_dir, 'grafico_simbolos.png')
    plt.tight_layout()
    plt.savefig(caminho, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    graficos.append(caminho)

    # ── Gráfico 3: Distribuição dos prêmios ──
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor('#1a1a2e')
    ax.set_facecolor('#16213e')

    contagem = resultados["contagem_premios"]
    if contagem:
        premios_sorted = sorted(contagem.keys())
        freqs = [contagem[p] for p in premios_sorted]
        labels = [f'{p:.0f}x' for p in premios_sorted]

        bars = ax.bar(labels, freqs, color='#00d4ff', edgecolor='#333', alpha=0.85)
        ax.set_xlabel('Multiplicador do Prêmio', color='white', fontsize=11)
        ax.set_ylabel('Frequência', color='white', fontsize=11)
        ax.set_title('Distribuição de Prêmios por Multiplicador', color='white', fontsize=13, fontweight='bold')
        ax.tick_params(colors='white', axis='both')
        ax.tick_params(axis='x', rotation=45)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        ax.grid(axis='y', color='#333', linestyle='--', alpha=0.5)

    caminho = os.path.join(output_dir, 'grafico_distribuicao.png')
    plt.tight_layout()
    plt.savefig(caminho, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    graficos.append(caminho)

    return graficos

# ─────────────────────────────────────────────
#  6. EXPORTAR PARA EXCEL
# ─────────────────────────────────────────────

def exportar_excel(resultados, rtp_teorico, detalhes_teorico, output_dir):
    """
    Gera o Game Math Document em Excel —
    formato padrão usado antes da certificação.
    """
    caminho = os.path.join(output_dir, f'{NOME_DO_JOGO.replace(" ","_")}_Math_Document.xlsx')
    wb = Workbook()

    # Estilos
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="00d4ff", bold=True, size=11)
    title_font  = Font(color="FFD700", bold=True, size=14)
    sub_font    = Font(color="FFFFFF", bold=True, size=11)
    data_font   = Font(color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin', color='444444'),
        right=Side(style='thin', color='444444'),
        top=Side(style='thin', color='444444'),
        bottom=Side(style='thin', color='444444')
    )
    dark_fill = PatternFill("solid", fgColor="16213e")
    mid_fill  = PatternFill("solid", fgColor="0f3460")

    def estilizar_header(ws, row, cols):
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    def estilizar_data(ws, row, cols, alt=False):
        fill = mid_fill if alt else dark_fill
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # ── Aba 1: Resumo ──
    ws1 = wb.active
    ws1.title = "📊 Resumo"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25

    ws1['A1'] = f"GAME MATH DOCUMENT — {NOME_DO_JOGO.upper()}"
    ws1['A1'].font = title_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:B1')

    dados_resumo = [
        ("Versão", VERSAO),
        ("Data", datetime.date.today().strftime("%d/%m/%Y")),
        ("", ""),
        ("CONFIGURAÇÃO DO JOGO", ""),
        ("Número de Rolos", NUM_ROLOS),
        ("Símbolos Únicos", len(SIMBOLOS)),
        ("RTP Alvo", f"{RTP_ALVO}%"),
        ("", ""),
        ("RESULTADOS DA SIMULAÇÃO", ""),
        ("Rodadas Simuladas", f"{resultados['num_rodadas']:,}"),
        ("Total Apostado", f"{resultados['total_apostado']:,.2f}"),
        ("Total Pago", f"{resultados['total_pago']:,.2f}"),
        ("RTP Teórico", f"{rtp_teorico:.4f}%"),
        ("RTP Simulado", f"{resultados['rtp_final']:.4f}%"),
        ("Diferença RTP", f"{abs(rtp_teorico - resultados['rtp_final']):.4f}%"),
        ("Hit Rate", f"{resultados['hit_rate']:.2f}%"),
    ]

    for i, (label, valor) in enumerate(dados_resumo, start=2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=valor)
        alt = i % 2 == 0
        estilizar_data(ws1, i, 2, alt)
        if label in ("CONFIGURAÇÃO DO JOGO", "RESULTADOS DA SIMULAÇÃO"):
            ws1.cell(row=i, column=1).font = sub_font
            ws1.merge_cells(f'A{i}:B{i}')

    # ── Aba 2: Tabela de Probabilidades ──
    ws2 = wb.create_sheet("🎰 Probabilidades")
    ws2.sheet_view.showGridLines = False

    headers = ["Símbolo", "Peso", "Prob. por Rolo (%)", "3x Mult", "4x Mult", "5x Mult"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    estilizar_header(ws2, 1, len(headers))

    total_pesos = sum(s["peso"] for s in SIMBOLOS.values())
    for i, (simbolo, dados) in enumerate(SIMBOLOS.items(), start=2):
        prob = dados["peso"] / total_pesos * 100
        pagtos = PAGAMENTOS.get(simbolo, {})
        row_data = [
            simbolo,
            dados["peso"],
            f"{prob:.2f}%",
            f'{pagtos.get(3,"-")}x' if pagtos.get(3) else "-",
            f'{pagtos.get(4,"-")}x' if pagtos.get(4) else "-",
            f'{pagtos.get(5,"-")}x' if pagtos.get(5) else "-",
        ]
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=col, value=val)
        estilizar_data(ws2, i, len(headers), i % 2 == 0)

    for col in range(1, len(headers)+1):
        ws2.column_dimensions[chr(64+col)].width = 18

    # ── Aba 3: RTP Detalhado ──
    ws3 = wb.create_sheet("📐 RTP Detalhado")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Símbolo", "Quantidade", "Multiplicador", "Probabilidade", "Prob (%)", "Contribuição RTP"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    estilizar_header(ws3, 1, len(headers3))

    for i, det in enumerate(detalhes_teorico, start=2):
        row_data = list(det.values())
        for col, val in enumerate(row_data, 1):
            ws3.cell(row=i, column=col, value=val)
        estilizar_data(ws3, i, len(headers3), i % 2 == 0)

    # Linha de total
    total_row = i + 2
    ws3.cell(row=total_row, column=1, value="TOTAL RTP TEÓRICO")
    ws3.cell(row=total_row, column=6, value=f"{rtp_teorico:.4f}%")
    for col in [1, 6]:
        ws3.cell(row=total_row, column=col).font = Font(color="FFD700", bold=True, size=11)
        ws3.cell(row=total_row, column=col).fill = header_fill

    for col in range(1, len(headers3)+1):
        ws3.column_dimensions[chr(64+col)].width = 20

    # ── Aba 4: Histórico de Simulação ──
    ws4 = wb.create_sheet("📈 Histórico RTP")
    ws4.sheet_view.showGridLines = False

    headers4 = ["Rodada", "RTP Simulado (%)", "RTP Teórico (%)", "Diferença (%)"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    estilizar_header(ws4, 1, len(headers4))

    for i, snap in enumerate(resultados["historico_rtp"], start=2):
        diff = abs(snap["rtp"] - rtp_teorico)
        row_data = [snap["rodada"], f"{snap['rtp']:.4f}", f"{rtp_teorico:.4f}", f"{diff:.4f}"]
        for col, val in enumerate(row_data, 1):
            ws4.cell(row=i, column=col, value=val)
        estilizar_data(ws4, i, len(headers4), i % 2 == 0)

    for col in range(1, len(headers4)+1):
        ws4.column_dimensions[chr(64+col)].width = 22

    wb.save(caminho)
    print(f"\n📊 Excel salvo: {caminho}")
    return caminho

# ─────────────────────────────────────────────
#  7. EXPORTAR PARA PDF
# ─────────────────────────────────────────────

def exportar_pdf(resultados, rtp_teorico, graficos, output_dir):
    """
    Gera o relatório técnico em PDF — formato usado
    para apresentação a operadores e certificadoras.
    """
    caminho = os.path.join(output_dir, f'{NOME_DO_JOGO.replace(" ","_")}_Report.pdf')
    doc = SimpleDocTemplate(caminho, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('title', fontSize=20, textColor=colors.HexColor('#FFD700'),
                                  spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
    style_h2 = ParagraphStyle('h2', fontSize=13, textColor=colors.HexColor('#00d4ff'),
                               spaceBefore=12, spaceAfter=6, fontName='Helvetica-Bold')
    style_body = ParagraphStyle('body', fontSize=10, textColor=colors.white,
                                 spaceAfter=4, fontName='Helvetica')
    style_caption = ParagraphStyle('caption', fontSize=9, textColor=colors.HexColor('#aaaaaa'),
                                    alignment=TA_CENTER, spaceAfter=8)

    bg_color = colors.HexColor('#1a1a2e')
    header_color = colors.HexColor('#0f3460')
    accent_color = colors.HexColor('#00d4ff')
    gold_color = colors.HexColor('#FFD700')
    text_color = colors.white
    alt_color = colors.HexColor('#16213e')

    elementos = []

    # Cabeçalho
    elementos.append(Spacer(1, 0.5*cm))
    elementos.append(Paragraph(f"GAME MATH DOCUMENT", style_title))
    elementos.append(Paragraph(f"{NOME_DO_JOGO} — v{VERSAO}", style_title))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph(f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", style_caption))
    elementos.append(Spacer(1, 0.5*cm))

    # Resumo executivo
    elementos.append(Paragraph("1. RESUMO EXECUTIVO", style_h2))

    diff = abs(rtp_teorico - resultados['rtp_final'])
    status = "✅ APROVADO" if diff < 0.5 else "⚠️ REVISAR"

    resumo_data = [
        ["PARÂMETRO", "VALOR"],
        ["RTP Alvo", f"{RTP_ALVO}%"],
        ["RTP Teórico Calculado", f"{rtp_teorico:.4f}%"],
        ["RTP Simulado (10M rodadas)", f"{resultados['rtp_final']:.4f}%"],
        ["Diferença Teórico vs Simulado", f"{diff:.4f}%"],
        ["Hit Rate", f"{resultados['hit_rate']:.2f}%"],
        ["Rodadas Simuladas", f"{resultados['num_rodadas']:,}"],
        ["Status de Validação", status],
    ]

    t = Table(resumo_data, colWidths=[9*cm, 7*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), accent_color),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [bg_color, alt_color]),
        ('TEXTCOLOR', (0,1), (-1,-1), text_color),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#444444')),
        ('ROWHEIGHT', (0,0), (-1,-1), 22),
        ('TEXTCOLOR', (1,-1), (1,-1), gold_color),
        ('FONTNAME', (1,-1), (1,-1), 'Helvetica-Bold'),
    ]))
    elementos.append(t)
    elementos.append(Spacer(1, 0.5*cm))

    # Tabela de pagamentos
    elementos.append(Paragraph("2. TABELA DE PAGAMENTOS", style_h2))

    total_pesos = sum(s["peso"] for s in SIMBOLOS.values())
    pag_data = [["Símbolo", "Prob./Rolo", "3 em linha", "4 em linha", "5 em linha"]]
    for simbolo, dados in SIMBOLOS.items():
        prob = dados["peso"] / total_pesos * 100
        pags = PAGAMENTOS.get(simbolo, {})
        pag_data.append([
            simbolo,
            f"{prob:.1f}%",
            f'{pags.get(3,"-")}x' if pags.get(3) else "-",
            f'{pags.get(4,"-")}x' if pags.get(4) else "-",
            f'{pags.get(5,"-")}x' if pags.get(5) else "-",
        ])

    t2 = Table(pag_data, colWidths=[5.5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), accent_color),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [bg_color, alt_color]),
        ('TEXTCOLOR', (0,1), (-1,-1), text_color),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#444444')),
        ('ROWHEIGHT', (0,0), (-1,-1), 20),
    ]))
    elementos.append(t2)
    elementos.append(Spacer(1, 0.5*cm))

    # Gráficos
    elementos.append(Paragraph("3. ANÁLISE GRÁFICA", style_h2))

    for i, grafico in enumerate(graficos):
        if os.path.exists(grafico):
            img = Image(grafico, width=16*cm, height=8*cm)
            elementos.append(img)
            captions = [
                "Figura 1 — Convergência do RTP ao longo das simulações (Lei dos Grandes Números)",
                "Figura 2 — Probabilidade de cada símbolo aparecer em um rolo",
                "Figura 3 — Distribuição de frequência dos prêmios por multiplicador",
            ]
            elementos.append(Paragraph(captions[i] if i < len(captions) else "", style_caption))
            elementos.append(Spacer(1, 0.3*cm))

    # Conclusão
    elementos.append(Paragraph("4. CONCLUSÃO", style_h2))
    conclusao = f"""
    O jogo <b>{NOME_DO_JOGO}</b> foi submetido à simulação Monte Carlo de <b>{resultados['num_rodadas']:,} rodadas</b>.
    O RTP teórico calculado matematicamente foi de <b>{rtp_teorico:.4f}%</b>, enquanto o RTP
    obtido na simulação foi de <b>{resultados['rtp_final']:.4f}%</b>, resultando em uma diferença
    de apenas <b>{diff:.4f}%</b> — dentro da margem aceitável para certificação (&lt;0.5%).
    O Hit Rate de <b>{resultados['hit_rate']:.2f}%</b> indica que o jogo paga algum prêmio em
    aproximadamente <b>1 a cada {100/resultados['hit_rate']:.0f} rodadas</b>.
    """
    elementos.append(Paragraph(conclusao, style_body))

    # Rodapé
    elementos.append(Spacer(1, 1*cm))
    elementos.append(Paragraph("─" * 80, style_caption))
    elementos.append(Paragraph("Documento gerado automaticamente pelo Slot Math Simulator — Portfólio iGaming", style_caption))

    # Build com fundo escuro
    def background(canvas, doc):
        canvas.setFillColor(bg_color)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)

    doc.build(elementos, onFirstPage=background, onLaterPages=background)
    print(f"📄 PDF salvo: {caminho}")
    return caminho

# ─────────────────────────────────────────────
#  8. MAIN — EXECUTA TUDO
# ─────────────────────────────────────────────

def main():
    print("\n" + "="*55)
    print(f"  🎰 SLOT MATH SIMULATOR — {NOME_DO_JOGO}")
    print("="*55)

    # Cria pasta de output
    output_dir = "/home/claude/slot_simulator/output"
    os.makedirs(output_dir, exist_ok=True)

    # 1. RTP Teórico
    print("\n📐 Calculando RTP teórico...")
    rtp_teorico, detalhes = calcular_rtp_teorico()
    print(f"   RTP Teórico: {rtp_teorico:.4f}%")

    # 2. Simulação
    resultados = simular(num_rodadas=10_000_000)

    # 3. Gráficos
    print("\n🎨 Gerando gráficos...")
    graficos = gerar_graficos(resultados, rtp_teorico, output_dir)

    # 4. Excel
    print("\n📊 Gerando planilha Excel...")
    exportar_excel(resultados, rtp_teorico, detalhes, output_dir)

    # 5. PDF
    print("\n📄 Gerando relatório PDF...")
    exportar_pdf(resultados, rtp_teorico, graficos, output_dir)

    # Resumo final
    print("\n" + "="*55)
    print("  ✅ PROJETO CONCLUÍDO!")
    print("="*55)
    print(f"\n  RTP Alvo:     {RTP_ALVO:.2f}%")
    print(f"  RTP Teórico:  {rtp_teorico:.4f}%")
    print(f"  RTP Simulado: {resultados['rtp_final']:.4f}%")
    diff = abs(rtp_teorico - resultados['rtp_final'])
    print(f"  Diferença:    {diff:.4f}%  {'✅ OK' if diff < 0.5 else '⚠️ Revisar'}")
    print(f"  Hit Rate:     {resultados['hit_rate']:.2f}%")
    print(f"\n  Arquivos gerados em: {output_dir}/")
    print("="*55 + "\n")

if __name__ == "__main__":
    main()
